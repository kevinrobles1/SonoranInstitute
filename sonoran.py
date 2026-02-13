from __future__ import annotations

import os
import re
import sys
import tempfile
import subprocess
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Optional, Any, Dict, List, Tuple

import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import load_workbook


# =============================
# App identity
# =============================
LOGO_URL = "https://sonoraninstitute.org/wp-content/themes/sonoran-institute-2016/assets/img/si_logo_2018.png"
PLOTLY_CONFIG = {"displaylogo": False}

st.set_page_config(page_title="Trash Dashboard", page_icon="🗑️", layout="wide")


# =============================
# Settings (env vars + defaults)
# =============================
@dataclass(frozen=True)
class Settings:
    input_xlsx: str
    output_xlsx: str
    overrides_csv: str
    clean_module: str
    sample_xlsx: str


def get_settings() -> Settings:
    base = Path(__file__).resolve().parent

    def env_or(key: str, default: str) -> str:
        v = os.getenv(key, "").strip()
        return v if v else default

    sample_xlsx = str(base / "data" / "sample_trash.xlsx")

    return Settings(
        input_xlsx=env_or("TRASH_INPUT_XLSX", sample_xlsx),
        output_xlsx=env_or("TRASH_OUTPUT_XLSX", str(base / "data" / "sample_trash_CLEANED.xlsx")),
        overrides_csv=env_or("TRASH_OVERRIDES_CSV", str(base / "data" / "site_overrides.csv")),
        clean_module=env_or("TRASH_CLEAN_MODULE", "scripts.clean_trash_db"),
        sample_xlsx=sample_xlsx,
    )


SETTINGS = get_settings()


# =============================
# UI helpers
# =============================
def inject_css_times_new_roman() -> None:
    st.markdown(
        """
        <style>
            html, body, [class*="css"] {
                font-family: "Times New Roman", Times, serif;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )


def st_df(df: pd.DataFrame, height: int = 420) -> None:
    try:
        st.dataframe(df, width="stretch", height=height)
    except TypeError:
        st.dataframe(df, use_container_width=True, height=height)


def st_chart(fig, height: Optional[int] = None) -> None:
    if height is not None:
        fig.update_layout(height=height)
    try:
        st.plotly_chart(fig, config=PLOTLY_CONFIG, width="stretch")
    except TypeError:
        st.plotly_chart(fig, config=PLOTLY_CONFIG, use_container_width=True)


def st_sidebar_logo(url: str) -> None:
    try:
        st.sidebar.image(url, width=170)
    except TypeError:
        st.sidebar.image(url, use_container_width=True)


# =============================
# General helpers
# =============================
def get_mtime(path: str) -> float:
    try:
        return os.path.getmtime(path)
    except Exception:
        return 0.0


def normalize_event_id(x: Any) -> Optional[str]:
    if pd.isna(x):
        return None
    s = str(x).strip()
    if s == "":
        return None
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s


def parse_date_val(val: Any) -> pd.Timestamp:
    if pd.isna(val):
        return pd.NaT

    if isinstance(val, (datetime, pd.Timestamp)):
        return pd.to_datetime(val, errors="coerce")

    s = str(val).strip()
    if s == "":
        return pd.NaT

    if re.fullmatch(r"\d+(\.0+)?", s):
        s = s.split(".")[0]

    if len(s) == 6:
        return pd.to_datetime(s, format="%y%m%d", errors="coerce")
    if len(s) == 8:
        return pd.to_datetime(s, format="%Y%m%d", errors="coerce")

    return pd.to_datetime(s, errors="coerce")


def pick_first_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols = set(df.columns)
    for c in candidates:
        if c in cols:
            return c
    return None


def multi_to_group_item(col: Any) -> Tuple[str, str]:
    if not isinstance(col, tuple) or len(col) != 2:
        return "Ungrouped", str(col)

    top, sub = col
    top = "" if pd.isna(top) else str(top).strip()
    sub = "" if pd.isna(sub) else str(sub).strip()

    group = top if (top and not top.lower().startswith("unnamed")) else "Ungrouped"
    item = sub if (sub and not sub.lower().startswith("unnamed")) else top

    if item == "":
        item = "Unknown item"

    return group, item


def looks_like_timedelta_text(x: Any) -> bool:
    if pd.isna(x):
        return False
    return ("day" in str(x)) or isinstance(x, pd.Timedelta)


def find_candidate_workbooks() -> List[str]:
    base = Path(__file__).resolve().parent
    candidates: List[str] = []

    obvious = [
        SETTINGS.output_xlsx,
        SETTINGS.input_xlsx,
        str(base / "Trash database_CLEANED.xlsx"),
        str(base / "Trash database.xlsx"),
        str(base / "data" / "Trash database.xlsx"),
        str(base / "data" / "Trash database_CLEANED.xlsx"),
        str(base / "data" / "sample_trash.xlsx"),
        str(base / "data" / "sample_trash_CLEANED.xlsx"),
    ]

    for p in obvious:
        if p and os.path.exists(p) and p not in candidates:
            candidates.append(p)

    for root in [base, base / "data"]:
        if root.is_dir():
            for fn in root.iterdir():
                if fn.is_file() and fn.name.lower().endswith(".xlsx"):
                    fp = str(fn)
                    if fp not in candidates:
                        candidates.append(fp)

    return candidates


# =============================
# Sample workbook creation
# =============================
def create_sample_if_missing() -> None:
    """
    Generates a fake workbook if data/sample_trash.xlsx does not exist.
    This keeps the repo runnable without any private data.
    """
    sample_path = Path(SETTINGS.sample_xlsx)
    if sample_path.exists():
        return

    try:
        from scripts.make_sample_workbook import create_sample_workbook  # type: ignore
    except Exception as e:
        st.error("Could not import scripts/make_sample_workbook.py")
        st.write(str(e))
        return

    sample_path.parent.mkdir(parents=True, exist_ok=True)
    create_sample_workbook(str(sample_path))


# =============================
# Plot-column builder for cleaned long table
# =============================
def build_plot_columns_for_clean_long(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    if "event_id" in out.columns:
        out["event_id"] = out["event_id"].apply(normalize_event_id)
    elif "Event ID" in out.columns:
        out["event_id"] = out["Event ID"].apply(normalize_event_id)
    else:
        out["event_id"] = None

    if "date_plot" not in out.columns:
        date_src = pick_first_col(out, ["date_resolved", "date_data", "date_site", "date", "Date"])
        out["date_plot"] = out[date_src].apply(parse_date_val) if date_src else pd.NaT
    else:
        out["date_plot"] = out["date_plot"].apply(parse_date_val)

    if "lat_plot" not in out.columns:
        lat_src = pick_first_col(out, ["lat", "lat_raw", "Latitude", "Lat"])
        out["lat_plot"] = pd.to_numeric(out[lat_src], errors="coerce") if lat_src else np.nan
    else:
        out["lat_plot"] = pd.to_numeric(out["lat_plot"], errors="coerce")

    if "lon_plot" not in out.columns:
        lon_src = pick_first_col(out, ["lon", "lon_raw", "Longitude", "Lon", "Long"])
        out["lon_plot"] = pd.to_numeric(out[lon_src], errors="coerce") if lon_src else np.nan
    else:
        out["lon_plot"] = pd.to_numeric(out["lon_plot"], errors="coerce")

    if "site_label_plot" not in out.columns:
        label_src = pick_first_col(out, ["site_label", "Site", "location_description"])
        out["site_label_plot"] = out[label_src].fillna("").astype(str).str.strip() if label_src else ""
        empty = out["site_label_plot"].eq("")
        out.loc[empty, "site_label_plot"] = out.loc[empty, "event_id"].apply(
            lambda x: f"unknown (event {x})" if x else "unknown"
        )
    else:
        out["site_label_plot"] = out["site_label_plot"].fillna("").astype(str).str.strip()

    if "count_for_totals" not in out.columns:
        if "count_clean" in out.columns:
            out["count_for_totals"] = pd.to_numeric(out["count_clean"], errors="coerce").fillna(0).astype(float)
        elif "count" in out.columns:
            out["count_for_totals"] = pd.to_numeric(out["count"], errors="coerce").fillna(0).astype(float)
        elif "count_raw" in out.columns:
            out["count_for_totals"] = pd.to_numeric(out["count_raw"], errors="coerce").fillna(0).astype(float)
        else:
            out["count_for_totals"] = 0.0
    else:
        out["count_for_totals"] = pd.to_numeric(out["count_for_totals"], errors="coerce").fillna(0).astype(float)

    return out


# =============================
# Workbook loader (raw or cleaned)
# =============================
@st.cache_data(show_spinner=False)
def load_workbook_any(workbook_path: str, workbook_mtime: float):
    xls = pd.ExcelFile(workbook_path)
    sheets = xls.sheet_names
    meta = {"sheets": sheets}

    extras: Dict[str, Any] = {
        "qc_report": None,
        "needs_fixes": None,
        "events_clean": None,
        "clean_long": None,
    }

    if ("Clean_Long" in sheets) and ("Events_clean" in sheets):
        mode = "cleaned"
        clean_long = pd.read_excel(workbook_path, sheet_name="Clean_Long")
        events_clean = pd.read_excel(workbook_path, sheet_name="Events_clean")

        extras["clean_long"] = clean_long
        extras["events_clean"] = events_clean

        if "QC_Report" in sheets:
            extras["qc_report"] = pd.read_excel(workbook_path, sheet_name="QC_Report")
        if "Needs_Fixes" in sheets:
            extras["needs_fixes"] = pd.read_excel(workbook_path, sheet_name="Needs_Fixes")

        long_df = build_plot_columns_for_clean_long(clean_long)

        if "trash_group" not in long_df.columns and "Trash group" in long_df.columns:
            long_df["trash_group"] = long_df["Trash group"]
        if "trash_item" not in long_df.columns and "Trash item" in long_df.columns:
            long_df["trash_item"] = long_df["Trash item"]

        events_df = long_df[["event_id", "date_plot"]].drop_duplicates().copy()
        if "surveyed_m2" in long_df.columns:
            tmp = long_df[["event_id", "surveyed_m2"]].drop_duplicates()
            events_df = events_df.merge(tmp, on="event_id", how="left")
        else:
            events_df["surveyed_m2"] = np.nan

        tmp2 = long_df[["event_id", "site_label_plot"]].drop_duplicates()
        events_df = events_df.merge(tmp2, on="event_id", how="left")

        site_df = None
        if "Site_clean" in sheets:
            try:
                site_df = pd.read_excel(workbook_path, sheet_name="Site_clean")
            except Exception:
                site_df = None

        return mode, events_df, long_df, site_df, meta, extras

    mode = "raw"

    if "Data" not in sheets:
        raise ValueError("Workbook missing 'Data' and missing cleaned sheets (Clean_Long, Events_clean).")

    df_raw = pd.read_excel(workbook_path, sheet_name="Data", header=[0, 1])

    base_targets = {
        "event id": "event_id",
        "date": "date_raw",
        "surveyed m2": "surveyed_m2_raw",
        "surveyed m^2": "surveyed_m2_raw",
    }

    base_cols: Dict[Any, str] = {}
    for col in df_raw.columns:
        sub = str(col[1]).strip().lower()
        if sub in base_targets:
            base_cols[col] = base_targets[sub]

    df = df_raw.copy()
    df.columns = [base_cols.get(c, c) for c in df_raw.columns]

    required = {"event_id", "date_raw", "surveyed_m2_raw"}
    missing = sorted(list(required - set(df.columns)))
    if missing:
        raise ValueError(f"Data sheet missing required columns: {missing}")

    df = df[df["event_id"].notna()].copy()

    df["event_id"] = df["event_id"].apply(normalize_event_id)
    df["date_plot"] = df["date_raw"].apply(parse_date_val)
    df["surveyed_m2"] = pd.to_numeric(df["surveyed_m2_raw"], errors="coerce")

    value_cols = [c for c in df.columns if isinstance(c, tuple)]
    long_df = df.melt(
        id_vars=["event_id", "date_plot", "surveyed_m2"],
        value_vars=value_cols,
        var_name="col",
        value_name="count_raw",
    )

    group_item = long_df["col"].apply(multi_to_group_item)
    long_df["trash_group"] = group_item.apply(lambda x: x[0])
    long_df["trash_item"] = group_item.apply(lambda x: x[1])

    drop_items = {"Complete?", "Total items", "Total items/m2"}
    long_df = long_df[~long_df["trash_item"].isin(drop_items)].copy()

    long_df["count_for_totals"] = pd.to_numeric(long_df["count_raw"], errors="coerce").fillna(0).astype(float)

    site_df = None
    if "Site" in sheets:
        site_df = pd.read_excel(workbook_path, sheet_name="Site")

        if "Event ID" in site_df.columns:
            site_df["event_id"] = site_df["Event ID"].apply(normalize_event_id)
        else:
            site_df["event_id"] = None

        if "Date" in site_df.columns:
            site_df["date_site"] = site_df["Date"].apply(parse_date_val)

        if "Site" not in site_df.columns:
            site_df["Site"] = None

        site_small = site_df[["event_id", "Site"]].drop_duplicates()
        long_df = long_df.merge(site_small, on="event_id", how="left")
        long_df["site_label_plot"] = long_df["Site"].fillna("").astype(str).str.strip()
    else:
        long_df["Site"] = None
        long_df["site_label_plot"] = ""

    empty = long_df["site_label_plot"].eq("")
    long_df.loc[empty, "site_label_plot"] = long_df.loc[empty, "event_id"].apply(
        lambda x: f"unknown (event {x})" if x else "unknown"
    )

    events_df = df[["event_id", "date_plot", "surveyed_m2"]].drop_duplicates().copy()
    if site_df is not None and "Site" in site_df.columns:
        events_df = events_df.merge(site_small, on="event_id", how="left")
        events_df = events_df.rename(columns={"Site": "site_label_plot"})
    else:
        events_df["site_label_plot"] = None

    return mode, events_df, long_df, site_df, meta, extras


# =============================
# Quality checks
# =============================
def run_data_quality_checks(events_df: pd.DataFrame, long_df: pd.DataFrame, site_df: Optional[pd.DataFrame]):
    issues: List[Dict[str, Any]] = []

    missing_event_id = events_df["event_id"].isna().sum()
    if missing_event_id > 0:
        issues.append(
            {"name": "Missing Event ID", "count": int(missing_event_id), "sample": events_df[events_df["event_id"].isna()].head(50)}
        )

    dupes = events_df["event_id"][events_df["event_id"].notna()].duplicated().sum()
    if dupes > 0:
        d = events_df[events_df["event_id"].duplicated(keep=False)].sort_values("event_id").head(100)
        issues.append({"name": "Duplicate Event ID", "count": int(dupes), "sample": d})

    date_bad = events_df["date_plot"].isna().sum()
    if date_bad > 0:
        issues.append({"name": "Dates that did not parse", "count": int(date_bad), "sample": events_df[events_df["date_plot"].isna()].head(50)})

    if "surveyed_m2" in events_df.columns:
        area_missing = events_df["surveyed_m2"].isna().sum()
        if area_missing > 0:
            issues.append({"name": "Missing surveyed m2", "count": int(area_missing), "sample": events_df[events_df["surveyed_m2"].isna()].head(50)})

        area_nonpos = (events_df["surveyed_m2"].notna() & (events_df["surveyed_m2"] <= 0)).sum()
        if area_nonpos > 0:
            issues.append(
                {
                    "name": "Surveyed m2 is 0 or negative",
                    "count": int(area_nonpos),
                    "sample": events_df[events_df["surveyed_m2"].notna() & (events_df["surveyed_m2"] <= 0)].head(50),
                }
            )

    if "count_raw" in long_df.columns:
        cnum = pd.to_numeric(long_df["count_raw"], errors="coerce")

        neg_counts = (cnum.notna() & (cnum < 0)).sum()
        if neg_counts > 0:
            issues.append({"name": "Negative trash counts", "count": int(neg_counts), "sample": long_df[cnum.notna() & (cnum < 0)].head(100)})

        nonint = (cnum.notna() & ((cnum % 1) != 0)).sum()
        if nonint > 0:
            issues.append({"name": "Trash counts that are not whole numbers", "count": int(nonint), "sample": long_df[cnum.notna() & ((cnum % 1) != 0)].head(100)})

    if site_df is not None and "event_id" in site_df.columns:
        data_ids = set(events_df["event_id"].dropna().tolist())
        site_ids = set(site_df["event_id"].dropna().tolist())

        missing_in_site = sorted(list(data_ids - site_ids))
        missing_in_data = sorted(list(site_ids - data_ids))

        if missing_in_site:
            issues.append({"name": "Event IDs in Data but missing in Site", "count": int(len(missing_in_site)), "sample": pd.DataFrame({"event_id": missing_in_site[:100]})})

        if missing_in_data:
            issues.append({"name": "Event IDs in Site but missing in Data", "count": int(len(missing_in_data)), "sample": pd.DataFrame({"event_id": missing_in_data[:100]})})

        if "Site" in site_df.columns:
            site_blank = site_df["Site"].isna().sum()
            if site_blank > 0:
                issues.append({"name": "Missing Site values in Site sheet", "count": int(site_blank), "sample": site_df[site_df["Site"].isna()].head(100)})

        for col in ["Northing", "Westing"]:
            if col in site_df.columns:
                bad = site_df[col].apply(looks_like_timedelta_text).sum()
                if bad > 0:
                    issues.append(
                        {
                            "name": f"{col} values that look time-like",
                            "count": int(bad),
                            "sample": site_df[site_df[col].apply(looks_like_timedelta_text)][["event_id", col]].head(100),
                        }
                    )

    return issues


# =============================
# Add Entry logic
# =============================
def yymmdd_int(d: date) -> int:
    return int(d.strftime("%y%m%d"))


def try_append_to_master(
    workbook_path: str,
    event_id: str,
    d: date,
    surveyed_m2: float,
    site_name: Optional[str],
    counts_rows: pd.DataFrame,
):
    wb = load_workbook(workbook_path)
    if "Data" not in wb.sheetnames:
        raise ValueError("Workbook has no 'Data' sheet. Add Entry works on a raw master workbook.")

    ws_data = wb["Data"]

    max_col = ws_data.max_column
    header_top = [ws_data.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    header_sub = [ws_data.cell(row=2, column=c).value for c in range(1, max_col + 1)]

    col_keys: List[Tuple[str, str]] = []
    for top, sub in zip(header_top, header_sub):
        top_s = "" if top is None else str(top).strip()
        sub_s = "" if sub is None else str(sub).strip()

        if sub_s.lower() in ["event id", "date", "surveyed m2", "surveyed m^2"]:
            col_keys.append(("__EVENT__", sub_s))
        else:
            g = top_s if (top_s and not top_s.lower().startswith("unnamed")) else "Ungrouped"
            i = sub_s if (sub_s and not sub_s.lower().startswith("unnamed")) else top_s
            col_keys.append((g, i))

    counts_rows = counts_rows.copy()
    counts_rows["trash_group"] = counts_rows["trash_group"].astype(str)
    counts_rows["trash_item"] = counts_rows["trash_item"].astype(str)
    counts_rows["count"] = pd.to_numeric(counts_rows["count"], errors="coerce").fillna(0)

    user_map: Dict[Tuple[str, str], float] = {}
    for _, r in counts_rows.iterrows():
        g = r["trash_group"].strip()
        i = r["trash_item"].strip()
        if g == "" or i == "":
            continue
        user_map[(g, i)] = float(r["count"])

    total_items = float(np.sum(list(user_map.values()))) if user_map else 0.0

    row_values: List[Any] = []
    for (g, i) in col_keys:
        if g == "__EVENT__":
            if i.lower() == "event id":
                row_values.append(event_id)
            elif i.lower() == "date":
                row_values.append(yymmdd_int(d))
            elif i.lower() in ["surveyed m2", "surveyed m^2"]:
                row_values.append(float(surveyed_m2) if surveyed_m2 is not None else None)
            else:
                row_values.append(None)
        else:
            if str(i).strip() == "Total items":
                row_values.append(total_items)
            elif str(i).strip() == "Total items/m2":
                if surveyed_m2 and surveyed_m2 > 0:
                    row_values.append(total_items / float(surveyed_m2))
                else:
                    row_values.append(None)
            else:
                row_values.append(user_map.get((g, i), None))

    ws_data.append(row_values)

    if "Site" in wb.sheetnames:
        ws_site = wb["Site"]
        site_headers = {str(ws_site.cell(row=1, column=c).value).strip(): c for c in range(1, ws_site.max_column + 1)}
        new_row = [None] * ws_site.max_column

        if "Event ID" in site_headers:
            new_row[site_headers["Event ID"] - 1] = event_id
        if "Date" in site_headers:
            new_row[site_headers["Date"] - 1] = yymmdd_int(d)
        if "Site" in site_headers:
            new_row[site_headers["Site"] - 1] = site_name

        ws_site.append(new_row)

    wb.save(workbook_path)


def write_to_staging(staging_path: str, event_row: dict, counts_rows: pd.DataFrame):
    with pd.ExcelWriter(staging_path, engine="openpyxl", mode="a" if os.path.exists(staging_path) else "w") as writer:
        ev = pd.DataFrame([event_row])
        ev.to_excel(writer, sheet_name="New_Events", index=False, header=not ("New_Events" in writer.book.sheetnames))

        cr = counts_rows.copy()
        cr["event_id"] = event_row["event_id"]
        cr.to_excel(writer, sheet_name="New_Counts", index=False, header=not ("New_Counts" in writer.book.sheetnames))


# =============================
# Cleaning runner
# =============================
def run_cleaning_module() -> Tuple[bool, str]:
    """
    Runs scripts/clean_trash_db.py as a module:
      python -m scripts.clean_trash_db
    Uses sys.executable so it runs in the same venv.
    """
    try:
        result = subprocess.run(
            [sys.executable, "-m", SETTINGS.clean_module],
            capture_output=True,
            text=True,
            cwd=str(Path(__file__).resolve().parent),
        )
        ok = (result.returncode == 0)
        out = (result.stdout or "") + "\n" + (result.stderr or "")
        return ok, out.strip()
    except Exception as e:
        return False, str(e)


# =============================
# Page rendering
# =============================
inject_css_times_new_roman()
st_sidebar_logo(LOGO_URL)
st.sidebar.markdown("### Trash Dashboard")

create_sample_if_missing()

candidates = find_candidate_workbooks()

uploaded = st.sidebar.file_uploader("Optional: open a different Excel file", type=["xlsx"])
manual_path = st.sidebar.text_input("Or paste a full file path", value="")
prefer_cleaned = st.sidebar.checkbox("Prefer cleaned workbook if available", value=True)

if "workbook_path" not in st.session_state:
    default_path = SETTINGS.output_xlsx if (prefer_cleaned and os.path.exists(SETTINGS.output_xlsx)) else SETTINGS.input_xlsx
    if not os.path.exists(default_path) and candidates:
        default_path = candidates[0]
    st.session_state["workbook_path"] = default_path if default_path else ""

workbook_path = st.session_state["workbook_path"]
temp_uploaded_path = None

if uploaded is not None:
    fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    with open(temp_path, "wb") as f:
        f.write(uploaded.getbuffer())
    workbook_path = temp_path
    temp_uploaded_path = temp_path
elif manual_path.strip() != "":
    workbook_path = manual_path.strip()
elif prefer_cleaned and os.path.exists(SETTINGS.output_xlsx):
    workbook_path = SETTINGS.output_xlsx
elif candidates:
    workbook_path = st.sidebar.selectbox("Workbook", candidates, index=0)

st.session_state["workbook_path"] = workbook_path

page = st.sidebar.radio(
    "Pages",
    ["Dashboard", "Figures", "Map", "Add Entry", "Data Quality Review", "Needs Fixes", "Cleaning", "Export"],
    index=0,
)

if not workbook_path or not os.path.exists(workbook_path):
    st.error("No workbook found. Use the sidebar to upload a file or set a valid path.")
    st.stop()

mtime = get_mtime(workbook_path)

try:
    mode, events_df, long_df, site_df, meta, extras = load_workbook_any(workbook_path, mtime)
except Exception as e:
    st.error(f"Could not load workbook: {e}")
    st.stop()

st.caption(f"Loaded: {workbook_path}")
st.caption(f"Mode: {mode} | Last modified: {datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')}")

if page == "Dashboard":
    st.title("Trash Dashboard")

    c1, c2, c3, c4 = st.columns(4)

    min_date = long_df["date_plot"].min()
    max_date = long_df["date_plot"].max()

    with c1:
        if pd.notna(min_date) and pd.notna(max_date):
            date_range = st.date_input("Date range", value=(min_date.date(), max_date.date()))
        else:
            date_range = st.date_input("Date range")

    with c2:
        sites = sorted([s for s in long_df["site_label_plot"].dropna().unique().tolist()])
        selected_sites = st.multiselect("Site", options=sites, default=sites)

    with c3:
        groups = sorted(long_df["trash_group"].dropna().unique().tolist()) if "trash_group" in long_df.columns else []
        selected_groups = st.multiselect("Trash group", options=groups, default=groups)

    with c4:
        items = sorted(long_df["trash_item"].dropna().unique().tolist()) if "trash_item" in long_df.columns else []
        selected_items = st.multiselect("Trash item", options=items, default=[])

    f = long_df.copy()

    if isinstance(date_range, (tuple, list)) and len(date_range) == 2:
        start_d, end_d = date_range
        f = f[f["date_plot"].notna()]
        f = f[(f["date_plot"].dt.date >= start_d) & (f["date_plot"].dt.date <= end_d)]

    if selected_sites:
        f = f[f["site_label_plot"].isin(selected_sites)]

    if selected_groups and "trash_group" in f.columns:
        f = f[f["trash_group"].isin(selected_groups)]

    if selected_items and "trash_item" in f.columns:
        f = f[f["trash_item"].isin(selected_items)]

    total_events = f["event_id"].nunique()
    total_items = float(f["count_for_totals"].sum())

    a, b, c = st.columns(3)
    a.metric("Events", int(total_events))
    b.metric("Total items (sum of counts)", int(total_items))
    c.metric("Rows shown", int(len(f)))

    st.divider()

    left, right = st.columns(2)

    with left:
        st.subheader("Totals over time (monthly)")
        ts = (
            f.dropna(subset=["date_plot"])
            .groupby(pd.Grouper(key="date_plot", freq="MS"))["count_for_totals"]
            .sum()
            .reset_index()
        )
        if len(ts) > 0:
            fig = px.line(ts, x="date_plot", y="count_for_totals")
            st_chart(fig, height=420)
        else:
            st.info("No rows for this filter.")

    with right:
        st.subheader("Top items")
        if "trash_item" in f.columns:
            top = (
                f.groupby(["trash_item"])["count_for_totals"]
                .sum()
                .sort_values(ascending=False)
                .head(15)
                .reset_index()
            )
            if len(top) > 0:
                fig = px.bar(top, x="count_for_totals", y="trash_item", orientation="h")
                st_chart(fig, height=420)
            else:
                st.info("No rows for this filter.")
        else:
            st.info("trash_item is missing in this workbook.")

    st.subheader("Filtered table")
    cols = [c for c in ["event_id", "date_plot", "site_label_plot", "trash_group", "trash_item", "count_for_totals", "surveyed_m2"] if c in f.columns]
    st_df(
        f[cols].sort_values([c for c in ["date_plot", "event_id"] if c in cols]),
        height=480,
    )

elif page == "Figures":
    st.title("Figures")

    figure_type = st.selectbox(
        "Choose a figure",
        [
            "Totals over time (monthly)",
            "Top items",
            "Top groups",
            "Site comparison (total items)",
            "Items per m2 (by site)",
        ],
        index=0,
    )

    df = long_df.copy()
    df["count_for_totals"] = pd.to_numeric(df["count_for_totals"], errors="coerce").fillna(0).astype(float)

    if figure_type == "Totals over time (monthly)":
        ts = (
            df.dropna(subset=["date_plot"])
            .groupby(pd.Grouper(key="date_plot", freq="MS"))["count_for_totals"]
            .sum()
            .reset_index()
        )
        fig = px.line(ts, x="date_plot", y="count_for_totals")
        st_chart(fig, height=520)

    elif figure_type == "Top items":
        if "trash_item" not in df.columns:
            st.error("trash_item is missing.")
        else:
            top = df.groupby("trash_item")["count_for_totals"].sum().sort_values(ascending=False).head(25).reset_index()
            fig = px.bar(top, x="count_for_totals", y="trash_item", orientation="h")
            st_chart(fig, height=720)

    elif figure_type == "Top groups":
        if "trash_group" not in df.columns:
            st.error("trash_group is missing.")
        else:
            top = df.groupby("trash_group")["count_for_totals"].sum().sort_values(ascending=False).reset_index()
            fig = px.bar(top, x="count_for_totals", y="trash_group", orientation="h")
            st_chart(fig, height=650)

    elif figure_type == "Site comparison (total items)":
        by_site = df.groupby("site_label_plot")["count_for_totals"].sum().sort_values(ascending=False).reset_index()
        fig = px.bar(by_site, x="count_for_totals", y="site_label_plot", orientation="h")
        st_chart(fig, height=760)

    elif figure_type == "Items per m2 (by site)":
        if "surveyed_m2" not in df.columns:
            st.error("surveyed_m2 is missing.")
        else:
            ev = df.groupby(["event_id", "site_label_plot", "surveyed_m2"], dropna=False)["count_for_totals"].sum().reset_index()
            ev["surveyed_m2"] = pd.to_numeric(ev["surveyed_m2"], errors="coerce")
            ev["items_per_m2"] = np.where(ev["surveyed_m2"] > 0, ev["count_for_totals"] / ev["surveyed_m2"], np.nan)
            agg = ev.groupby("site_label_plot")["items_per_m2"].mean().sort_values(ascending=False).reset_index()
            fig = px.bar(agg, x="items_per_m2", y="site_label_plot", orientation="h")
            st_chart(fig, height=760)

elif page == "Map":
    st.title("Map")

    m = long_df.copy()
    if ("lat_plot" not in m.columns) or ("lon_plot" not in m.columns):
        st.error("Missing lat_plot or lon_plot. Run Cleaning or load a workbook that has coordinates.")
        st.stop()

    ev = (
        m.groupby(["event_id", "site_label_plot", "date_plot", "lat_plot", "lon_plot"], dropna=False)["count_for_totals"]
        .sum()
        .reset_index()
    )
    ev = ev[ev["lat_plot"].notna() & ev["lon_plot"].notna()].copy()

    if len(ev) == 0:
        st.warning("No events have usable coordinates.")
        st.stop()

    fig = px.scatter_mapbox(
        ev,
        lat="lat_plot",
        lon="lon_plot",
        hover_name="site_label_plot",
        hover_data={"event_id": True, "date_plot": True, "count_for_totals": True},
        zoom=7,
        height=700,
    )
    fig.update_layout(mapbox_style="open-street-map")
    st_chart(fig)

elif page == "Data Quality Review":
    st.title("Data Quality Review")

    if extras.get("qc_report") is not None and len(extras["qc_report"]) > 0:
        st.subheader("QC_Report (from workbook)")
        st_df(extras["qc_report"], height=420)
        st.divider()

    st.write("These checks are computed from the workbook loaded right now.")
    issues = run_data_quality_checks(events_df, long_df, site_df)

    if not issues:
        st.success("No issues found by the checks that are turned on.")
    else:
        st.warning(f"Issues found: {len(issues)}")
        for issue in issues:
            st.subheader(f"{issue['name']} (count: {issue['count']})")
            st_df(issue["sample"], height=280)

elif page == "Needs Fixes":
    st.title("Needs Fixes")

    nf = extras.get("needs_fixes")
    if nf is None:
        st.info("Needs_Fixes sheet not found in this workbook.")
    elif len(nf) == 0:
        st.success("Needs_Fixes is empty.")
    else:
        st.warning("Needs_Fixes is not empty.")
        st_df(nf, height=520)
        if "event_id" in nf.columns:
            ids = nf["event_id"].dropna().astype(str).tolist()
            st.write(f"Event IDs: {ids[:200]}")
        st.write("Overrides CSV path:")
        st.code(SETTINGS.overrides_csv)

elif page == "Add Entry":
    st.title("Add Entry")

    st.write("This page appends to a raw master workbook that has a 'Data' sheet.")
    st.write("If the file is locked, it writes to a staging file instead.")

    group_choices = sorted(long_df["trash_group"].dropna().unique().tolist()) if "trash_group" in long_df.columns else ["Ungrouped"]
    item_choices = sorted(long_df["trash_item"].dropna().unique().tolist()) if "trash_item" in long_df.columns else [""]

    with st.form("add_entry_form"):
        c1, c2, c3 = st.columns(3)
        with c1:
            event_id = st.text_input("Event ID")
        with c2:
            entry_date = st.date_input("Date")
        with c3:
            surveyed_m2 = st.number_input("Surveyed m2", min_value=0.0, value=0.0, step=0.5)

        site_name = st.text_input("Site (optional)")

        st.markdown("#### Trash counts")
        starter = pd.DataFrame(
            [{
                "trash_group": (group_choices[0] if group_choices else "Ungrouped"),
                "trash_item": (item_choices[0] if item_choices else ""),
                "count": 0
            }]
        )

        try:
            counts_rows = st.data_editor(
                starter,
                num_rows="dynamic",
                width="stretch",
                column_config={
                    "trash_group": st.column_config.SelectboxColumn("Trash group", options=group_choices),
                    "trash_item": st.column_config.SelectboxColumn("Trash item", options=item_choices),
                    "count": st.column_config.NumberColumn("Count", min_value=0),
                },
            )
        except TypeError:
            counts_rows = st.data_editor(starter, num_rows="dynamic", use_container_width=True)

        block_if_duplicate = st.checkbox("Block duplicate Event ID", value=True)
        submit = st.form_submit_button("Save")

    if submit:
        event_id_norm = normalize_event_id(event_id)
        if not event_id_norm:
            st.error("Event ID is required.")
            st.stop()

        already = set(events_df["event_id"].dropna().tolist())
        if block_if_duplicate and event_id_norm in already:
            st.error("That Event ID already exists in the loaded workbook.")
            st.stop()

        if surveyed_m2 is None or surveyed_m2 <= 0:
            st.warning("Surveyed m2 is 0 or less. Items per m2 will be blank for this event.")

        counts_df = pd.DataFrame(counts_rows)
        if len(counts_df) == 0:
            counts_df = pd.DataFrame(columns=["trash_group", "trash_item", "count"])

        counts_df["count"] = pd.to_numeric(counts_df["count"], errors="coerce").fillna(0)
        counts_df = counts_df[(counts_df["trash_group"].notna()) & (counts_df["trash_item"].notna())]

        target_path = st.session_state.get("master_write_path", SETTINGS.input_xlsx)
        target_path = st.text_input("Master workbook path to write into", value=target_path)
        st.session_state["master_write_path"] = target_path

        if not os.path.exists(target_path):
            st.error("Master workbook path does not exist.")
            st.stop()

        try:
            try_append_to_master(
                workbook_path=target_path,
                event_id=event_id_norm,
                d=entry_date,
                surveyed_m2=float(surveyed_m2),
                site_name=site_name if site_name.strip() != "" else None,
                counts_rows=counts_df,
            )
            st.success("Saved to the master workbook.")
            st.cache_data.clear()
        except Exception as e:
            staging_path = str(Path(__file__).resolve().parent / "staging_new_entries.xlsx")
            event_row = {
                "event_id": event_id_norm,
                "date": entry_date.isoformat(),
                "surveyed_m2": float(surveyed_m2),
                "site": site_name if site_name.strip() != "" else None,
                "error": str(e),
            }
            try:
                write_to_staging(staging_path, event_row, counts_df)
                st.error("Could not write to the master workbook (it may be locked).")
                st.info(f"Saved to staging file instead: {staging_path}")
            except Exception as e2:
                st.error("Could not write to the master workbook, and staging write also failed.")
                st.write(str(e))
                st.write(str(e2))

elif page == "Cleaning":
    st.title("Cleaning")

    st.write("This runs the cleaning script and writes a cleaned workbook with Clean_Long and related sheets.")
    st.write("It runs: python -m scripts.clean_trash_db")

    st.code(
        "\n".join(
            [
                f"Input:    {SETTINGS.input_xlsx}",
                f"Output:   {SETTINGS.output_xlsx}",
                f"Overrides:{SETTINGS.overrides_csv}",
                f"Module:   {SETTINGS.clean_module}",
            ]
        )
    )

    run_now = st.button("Run cleaning now", type="primary")
    if run_now:
        with st.spinner("Running cleaning..."):
            ok, output = run_cleaning_module()

        if ok:
            st.success("Cleaning finished.")
            if output:
                st.code(output)
            st.cache_data.clear()

            if os.path.exists(SETTINGS.output_xlsx):
                st.session_state["workbook_path"] = SETTINGS.output_xlsx
                st.info("Switched workbook to the cleaned output.")
        else:
            st.error("Cleaning failed
